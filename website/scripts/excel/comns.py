import sqlite3
from openpyxl import load_workbook

db_path = '../../../db.sqlite3'


def script_connect():
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    return conn, c


def input_excel_pianoboek_peters():
    inp = '/Users/orion/PycharmProjects/rename/output/scarlatti/Scarlatti Sonaten edPeters.xlsx'
    wb = load_workbook(inp)
    ws = wb.get_sheet_by_name('Scarlatti Sonaten edPeters')
    pianoboek = {
        'band I': 1,
        'band II': 2,
        'band III': 3,
        '24 sonaten': 4,
        'One Volume': 5,
        'classic edition': 10,
    }
    conn, c = script_connect()
    sql = '''
    INSERT OR IGNORE INTO LibraryCode_Pianoboek 
    (LibraryCode, PianoboekID, Nr) 
    VALUES (?,?,?)
    '''
    for cellObj in ws['A1':'D257']:
        data = []
        for cells in cellObj:
            data.append(cells.value)
        if data[3] > 0 and data[3] != 'k':
            pianoboek_id = pianoboek.get(data[0])
            if pianoboek_id:
                code = 'K {}'.format(data[3])
                c.execute(sql, (code, pianoboek_id, data[1]))
                conn.commit()
            else:
                print data


def input_excel_pianoboek_henle():
    inp = '/Users/orion/PycharmProjects/rename/output/scarlatti/Scarlatti Sonaten Henle.xlsx'
    wb = load_workbook(inp)
    ws = wb.get_sheet_by_name('Scarlatti Sonaten Henle')
    pianoboek = {
        'band I': 6,
        'band II': 7,
        'band III': 8,
        'band IV': 9,
    }
    conn, c = script_connect()
    sql = '''
    INSERT OR IGNORE INTO LibraryCode_Pianoboek 
    (LibraryCode, PianoboekID, Nr) 
    VALUES (?,?,?)
    '''
    for cellObj in ws['A5':'D104']:
        data = []
        for cells in cellObj:
            data.append(cells.value)
        # print data
        if data[3] > 0 and data[3] != 'k':
            pianoboek_id = pianoboek.get(data[0])
            if pianoboek_id:
                code = 'K {}'.format(data[3])
                c.execute(sql, (code, pianoboek_id, data[1]))
                conn.commit()
            else:
                print data


def input_excel_tempo():
    inp = '/Users/orion/PycharmProjects/rename/output/scarlatti/Scarlatti Sonaten edPeters.xlsx'
    wb = load_workbook(inp)
    ws = wb.get_sheet_by_name('Scarlatti Sonaten edPeters')
    conn, c = script_connect()
    sql = '''
    INSERT OR IGNORE INTO LibraryCode (Code, Tempo, Key) 
    VALUES (?,?,?)
    '''
    for cellObj in ws['A1':'E257']:
        data = []
        for cells in cellObj:
            data.append(cells.value)
        if data[3] > 0 and data[3] != 'k':
            code = 'K {}'.format(data[3])
            c.execute(sql, (code, data[4], data[2]))
            conn.commit()


def main():
    # input_excel_pianoboek_peters()
    # input_excel_tempo()
    input_excel_pianoboek_henle()


if __name__ == '__main__':
    main()
